"""
Extract Australian occupation data from Jobs and Skills Australia Excel files.

Downloads and processes:
- Occupation Profiles (employment, earnings, education, descriptions, tasks)
- Employment Projections (5-year and 10-year growth outlook)

Data sources:
- Jobs and Skills Australia, Occupation Profiles Data (Feb 2025)
  https://www.jobsandskills.gov.au/data/occupation-and-industry-profiles
- Jobs and Skills Australia, Employment Projections May 2024 – May 2034
  https://www.jobsandskills.gov.au/data/employment-projections

Produces:
- occupations.json  — master list with metadata
- occupations.csv   — structured stats for all occupations
- pages/<slug>.md   — markdown descriptions for AI scoring

Usage:
    uv run python extract_data.py
    uv run python extract_data.py --no-download   # skip download, use cached files
"""

import argparse
import csv
import json
import os
import re
import urllib.request

import openpyxl


# ---------------------------------------------------------------------------
# Data source URLs (official Jobs and Skills Australia)
# ---------------------------------------------------------------------------
PROFILES_URL = (
    "https://www.jobsandskills.gov.au/sites/default/files/2025-04/"
    "occupation_profiles_data_-_february_2025.xlsx"
)
PROJECTIONS_URL = (
    "https://www.jobsandskills.gov.au/sites/default/files/2024-11/"
    "employment_projections_-_may_2024_to_may_2034.xlsx"
)

PROFILES_FILE = "occupation_profiles.xlsx"
PROJECTIONS_FILE = "employment_projections.xlsx"

# ANZSCO 1-digit major groups
ANZSCO_MAJOR_GROUPS = {
    1: "Managers",
    2: "Professionals",
    3: "Technicians and Trades Workers",
    4: "Community and Personal Service Workers",
    5: "Clerical and Administrative Workers",
    6: "Sales Workers",
    7: "Machinery Operators and Drivers",
    8: "Labourers",
}


def slugify(title):
    """Convert occupation title to a URL-friendly slug."""
    s = title.lower()
    s = re.sub(r"[^a-z0-9\s-]", "", s)
    s = re.sub(r"[\s]+", "-", s).strip("-")
    return s


def download_files(force=False):
    """Download Excel files from JSA if not already cached."""
    for url, path in [(PROFILES_URL, PROFILES_FILE), (PROJECTIONS_URL, PROJECTIONS_FILE)]:
        if not force and os.path.exists(path):
            print(f"  CACHED {path}")
            continue
        print(f"  Downloading {path}...", end=" ", flush=True)
        urllib.request.urlretrieve(url, path)
        size = os.path.getsize(path) / 1024
        print(f"OK ({size:.0f} KB)")


def load_profiles():
    """Load occupation profiles data from the JSA Excel file."""
    wb = openpyxl.load_workbook(PROFILES_FILE, read_only=True)

    # Table 1 - Overview: ANZSCO, title, employed, part-time%, female%, median weekly earnings, median age, annual growth
    overview = {}
    ws = wb["Table_1"]
    for row in ws.iter_rows(min_row=8, values_only=True):
        code = row[0]
        if not isinstance(code, (int, float)):
            continue
        code = int(code)
        overview[code] = {
            "anzsco": code,
            "title": row[1],
            "employed": row[2],
            "part_time_pct": row[3],
            "female_pct": row[4],
            "median_weekly_earnings": row[5] if row[5] != "N/A" else None,
            "median_age": row[6],
            "annual_growth": row[7],
        }

    # Table 2 - Descriptions
    descriptions = {}
    ws = wb["Table_2"]
    for row in ws.iter_rows(min_row=8, values_only=True):
        code = row[0]
        if not isinstance(code, (int, float)):
            continue
        descriptions[int(code)] = row[2]

    # Table 3 - Tasks (multiple rows per occupation)
    tasks = {}
    ws = wb["Table_3"]
    for row in ws.iter_rows(min_row=8, values_only=True):
        code = row[0]
        if not isinstance(code, (int, float)):
            continue
        code = int(code)
        if code not in tasks:
            tasks[code] = []
        if row[2]:
            tasks[code].append(row[2])

    # Table 4 - Earnings and hours
    earnings = {}
    ws = wb["Table_4"]
    for row in ws.iter_rows(min_row=8, values_only=True):
        code = row[0]
        if not isinstance(code, (int, float)):
            continue
        code = int(code)
        earnings[code] = {
            "ft_share_pct": row[2],
            "avg_ft_hours": row[3],
            "median_ft_weekly": row[4] if row[4] != "N/A" else None,
            "median_ft_hourly": row[5] if row[5] != "N/A" else None,
        }

    # Table 8 - Education
    education = {}
    ws = wb["Table_8"]
    for row in ws.iter_rows(min_row=8, values_only=True):
        code = row[0]
        if not isinstance(code, (int, float)):
            continue
        code = int(code)
        education[code] = {
            "postgrad_pct": row[2],
            "bachelor_pct": row[3],
            "adv_diploma_pct": row[4],
            "cert_iii_iv_pct": row[5],
            "year_12_pct": row[6],
            "year_11_pct": row[7],
            "year_10_below_pct": row[8],
        }

    wb.close()
    return overview, descriptions, tasks, earnings, education


def load_projections():
    """Load employment projections from the JSA Excel file."""
    wb = openpyxl.load_workbook(PROJECTIONS_FILE, read_only=True)
    ws = wb["Table_6 Occupation Unit Group"]

    projections = {}
    for row in ws.iter_rows(min_row=10, values_only=True):
        occ_level = row[0]
        nfd = row[1]
        code = row[2]

        # Only 4-digit ANZSCO, non-NFD
        if occ_level != 4 or nfd != "N":
            continue
        if not isinstance(code, (int, float)):
            continue

        code = int(code)
        projections[code] = {
            "skill_level": row[4],
            "baseline_thousands": row[5],
            "projected_2029_thousands": row[6],
            "projected_2034_thousands": row[7],
            "change_5yr_thousands": row[8],
            "change_5yr_pct": row[9],
            "change_10yr_thousands": row[10],
            "change_10yr_pct": row[11],
        }

    wb.close()
    return projections


def dominant_education(edu_data):
    """Determine the most common education level for an occupation."""
    if not edu_data:
        return ""

    levels = [
        ("Postgraduate degree", edu_data.get("postgrad_pct", 0) or 0),
        ("Bachelor degree", edu_data.get("bachelor_pct", 0) or 0),
        ("Advanced Diploma/Diploma", edu_data.get("adv_diploma_pct", 0) or 0),
        ("Certificate III/IV", edu_data.get("cert_iii_iv_pct", 0) or 0),
        ("Year 12", edu_data.get("year_12_pct", 0) or 0),
        ("Year 11 or below", (edu_data.get("year_11_pct", 0) or 0) + (edu_data.get("year_10_below_pct", 0) or 0)),
    ]
    return max(levels, key=lambda x: x[1])[0]


def outlook_description(pct):
    """Convert growth percentage to a descriptive label (Australian context)."""
    if pct is None:
        return ""
    if pct >= 15:
        return "Very strong growth"
    elif pct >= 10:
        return "Strong growth"
    elif pct >= 5:
        return "Moderate growth"
    elif pct >= 0:
        return "Stable"
    else:
        return "Declining"


def main():
    parser = argparse.ArgumentParser(description="Extract Australian occupation data")
    parser.add_argument("--no-download", action="store_true", help="Skip downloading, use cached files")
    parser.add_argument("--force-download", action="store_true", help="Re-download even if cached")
    parser.add_argument("--anzsco-digits", type=int, default=4, choices=[4, 6],
                        help="ANZSCO level: 4-digit (358 unit groups) or 6-digit (1236 occupations)")
    args = parser.parse_args()

    # Step 1: Download
    if not args.no_download:
        print("Downloading data from Jobs and Skills Australia...")
        download_files(force=args.force_download)

    # Step 2: Load all data
    print("\nLoading occupation profiles...")
    overview, descriptions, tasks, earnings, education = load_profiles()
    print(f"  {len(overview)} occupations in profiles")

    print("Loading employment projections...")
    projections = load_projections()
    print(f"  {len(projections)} occupations in projections")

    # Step 3: Merge into unified records
    # Filter by ANZSCO digit level
    target_digits = args.anzsco_digits
    filtered_overview = {
        code: info for code, info in overview.items()
        if len(str(code)) == target_digits
    }
    print(f"\nFiltered to {target_digits}-digit ANZSCO: {len(filtered_overview)} occupations")

    occupations = []
    for code, info in sorted(filtered_overview.items()):
        title = info["title"]
        slug = slugify(title)
        major_group = int(str(code)[0])
        category = ANZSCO_MAJOR_GROUPS.get(major_group, "Other")
        category_slug = slugify(category)

        proj = projections.get(code, {})
        earn = earnings.get(code, {})
        edu = education.get(code, {})

        # Compute median annual pay from weekly earnings
        median_weekly = info.get("median_weekly_earnings")
        if isinstance(median_weekly, (int, float)):
            median_annual = int(median_weekly * 52)
        else:
            # Try full-time weekly from Table 4
            ft_weekly = earn.get("median_ft_weekly")
            if isinstance(ft_weekly, (int, float)):
                median_annual = int(ft_weekly * 52)
            else:
                median_annual = None

        # Employment count (use profiles data in persons, not thousands)
        employed = info.get("employed")
        if isinstance(employed, (int, float)):
            employed = int(employed)
        else:
            employed = None

        # Growth outlook (5-year percentage)
        outlook_pct = proj.get("change_5yr_pct")
        if isinstance(outlook_pct, (int, float)):
            outlook_pct_display = round(outlook_pct * 100, 1)
        else:
            outlook_pct_display = None

        occ = {
            "title": title,
            "slug": slug,
            "anzsco": str(code),
            "category": category_slug,
            "category_name": category,
            "url": f"https://www.jobsandskills.gov.au/data/occupation-and-industry-profiles/occupations/{code}-{slug}",
            "median_pay_annual": median_annual,
            "median_pay_weekly": int(median_weekly) if isinstance(median_weekly, (int, float)) else None,
            "median_ft_hourly": earn.get("median_ft_hourly") if isinstance(earn.get("median_ft_hourly"), (int, float)) else None,
            "entry_education": dominant_education(edu),
            "employed": employed,
            "outlook_pct": outlook_pct_display,
            "outlook_desc": outlook_description(outlook_pct_display) if outlook_pct_display is not None else "",
            "projected_2029": int(proj["projected_2029_thousands"] * 1000) if proj.get("projected_2029_thousands") else None,
            "projected_2034": int(proj["projected_2034_thousands"] * 1000) if proj.get("projected_2034_thousands") else None,
            "skill_level": proj.get("skill_level"),
            "part_time_pct": info.get("part_time_pct"),
            "female_pct": info.get("female_pct"),
            "median_age": info.get("median_age"),
            "description": descriptions.get(code, ""),
            "tasks": tasks.get(code, []),
        }
        occupations.append(occ)

    print(f"\nMerged {len(occupations)} occupations")

    # Step 4: Write occupations.json (master list for other scripts)
    occ_list = [
        {
            "title": o["title"],
            "slug": o["slug"],
            "anzsco": o["anzsco"],
            "category": o["category"],
            "category_name": o["category_name"],
            "url": o["url"],
        }
        for o in occupations
    ]
    with open("occupations.json", "w") as f:
        json.dump(occ_list, f, indent=2)
    print(f"Wrote {len(occ_list)} occupations to occupations.json")

    # Step 5: Write occupations.csv
    fieldnames = [
        "title", "category", "category_name", "slug", "anzsco",
        "median_pay_annual", "median_pay_weekly", "median_ft_hourly",
        "entry_education", "skill_level",
        "employed", "projected_2029", "projected_2034",
        "outlook_pct", "outlook_desc",
        "part_time_pct", "female_pct", "median_age",
        "url",
    ]
    with open("occupations.csv", "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for o in occupations:
            row = {k: o.get(k, "") for k in fieldnames}
            # Convert None to empty string
            for k in row:
                if row[k] is None:
                    row[k] = ""
            writer.writerow(row)
    print(f"Wrote {len(occupations)} rows to occupations.csv")

    # Step 6: Generate Markdown pages for AI scoring
    os.makedirs("pages", exist_ok=True)
    pages_written = 0
    for o in occupations:
        md = []
        md.append(f"# {o['title']}")
        md.append("")
        md.append(f"**ANZSCO Code:** {o['anzsco']}")
        md.append(f"**Category:** {o['category_name']}")
        md.append(f"**Source:** [Jobs and Skills Australia]({o['url']})")
        md.append("")

        if o["description"]:
            md.append("## Description")
            md.append("")
            md.append(o["description"])
            md.append("")

        if o["tasks"]:
            md.append("## Key Tasks")
            md.append("")
            for task in o["tasks"]:
                md.append(f"- {task}")
            md.append("")

        md.append("## Quick Facts")
        md.append("")
        md.append("| Field | Value |")
        md.append("|-------|-------|")
        if o["median_pay_annual"]:
            md.append(f"| Median Annual Pay | ${o['median_pay_annual']:,} |")
        if o["median_pay_weekly"]:
            md.append(f"| Median Weekly Earnings | ${o['median_pay_weekly']:,} |")
        if o["median_ft_hourly"]:
            md.append(f"| Median Hourly (Full-time) | ${o['median_ft_hourly']:.2f} |")
        if o["employed"]:
            md.append(f"| Employment | {o['employed']:,} |")
        if o["entry_education"]:
            md.append(f"| Typical Education | {o['entry_education']} |")
        if o["skill_level"]:
            md.append(f"| Skill Level | {o['skill_level']} |")
        if o["outlook_pct"] is not None:
            md.append(f"| 5-Year Growth Outlook | {o['outlook_pct']}% ({o['outlook_desc']}) |")
        if o["part_time_pct"] is not None:
            md.append(f"| Part-time Share | {o['part_time_pct']}% |")
        if o["female_pct"] is not None:
            md.append(f"| Female Share | {o['female_pct']}% |")
        if o["median_age"] is not None:
            md.append(f"| Median Age | {o['median_age']} |")
        md.append("")

        md_path = f"pages/{o['slug']}.md"
        with open(md_path, "w") as f:
            f.write("\n".join(md))
        pages_written += 1

    print(f"Wrote {pages_written} Markdown pages to pages/")

    # Summary stats
    with_pay = sum(1 for o in occupations if o["median_pay_annual"])
    with_jobs = sum(1 for o in occupations if o["employed"])
    with_outlook = sum(1 for o in occupations if o["outlook_pct"] is not None)
    total_employed = sum(o["employed"] for o in occupations if o["employed"])

    print(f"\nSummary:")
    print(f"  Total occupations: {len(occupations)}")
    print(f"  With pay data: {with_pay}")
    print(f"  With employment data: {with_jobs}")
    print(f"  With outlook data: {with_outlook}")
    print(f"  Total employment represented: {total_employed:,}")

    # Sample
    print(f"\nSample occupations:")
    for o in occupations[:5]:
        pay = f"${o['median_pay_annual']:,}/yr" if o['median_pay_annual'] else "N/A"
        jobs = f"{o['employed']:,}" if o['employed'] else "N/A"
        outlook = f"{o['outlook_pct']}%" if o['outlook_pct'] is not None else "N/A"
        print(f"  {o['title']}: {pay}, {jobs} jobs, {outlook} outlook")


if __name__ == "__main__":
    main()
